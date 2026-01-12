from typing import Any, Dict, List
from io import BytesIO
from app.schemas.layout import LayoutConfig, LayoutSection, LayoutField, LayoutColumn

class DynamicGenerator:
    """
    Generates HTML, PDF, and Excel based on a LayoutConfig and extracted EDI data.
    """
    
    def __init__(self, config: LayoutConfig):
        self.config = config

    def generate(self, doc) -> bytes:
        """
        Main entry point for HTML. Generates the full HTML document.
        """
        # 1. Build Document Title - support custom placeholders from header
        try:
            format_data = {
                'name': doc.transaction_name,
                'ref_number': doc.header.get('po_number', 'Unknown'),
                **doc.header  # Pass all header fields for custom placeholders
            }
            title = self.config.title_format.format(**format_data)
        except KeyError:
            title = f"{doc.transaction_name}"
        
        # 2. Render Content
        content_html = self.render_content(doc)
        
        # 3. Wrap in Main Template (Premium Style)
        full_html = f"""
        <div class="main-container">
            <!-- Header Card -->
            <div class="nav-card">
                <h1>{title}</h1>
                <div class="nav-links">
                    <span class="active">View</span>
                    <span>Raw Data</span>
                </div>
            </div>

            <!-- Dynamic Content -->
            {content_html}
        </div>
        """
        
        return self._wrap_in_html(title, full_html)

    def generate_pdf(self, documents: list) -> bytes:
        """
        Generate PDF from documents using layout config.
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        buffer = BytesIO()
        doc_pdf = SimpleDocTemplate(buffer, pagesize=letter, 
                                    leftMargin=0.5*inch, rightMargin=0.5*inch,
                                    topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=18, spaceAfter=12, textColor=colors.HexColor('#1e40af'))
        section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'],
                                       fontSize=12, spaceBefore=16, spaceAfter=8,
                                       textColor=colors.HexColor('#1e40af'))
        label_style = ParagraphStyle('Label', parent=styles['Normal'],
                                     fontSize=9, textColor=colors.HexColor('#64748b'))
        value_style = ParagraphStyle('Value', parent=styles['Normal'],
                                     fontSize=10, textColor=colors.HexColor('#0f172a'))
        
        elements = []
        
        for i, edi_doc in enumerate(documents):
            if i > 0:
                elements.append(PageBreak())
            
            # Title - support custom placeholders from header
            try:
                format_data = {
                    'name': edi_doc.transaction_name,
                    'ref_number': edi_doc.header.get('po_number') or edi_doc.header.get('credit_debit_number', 'Unknown'),
                    **edi_doc.header
                }
                title = self.config.title_format.format(**format_data)
            except KeyError:
                title = f"{edi_doc.transaction_name}"
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 12))
            
            # Render sections based on config
            for section in self.config.sections:
                if not section.visible:
                    continue
                
                elements.append(Paragraph(section.title, section_style))
                
                if section.type == "fields":
                    # Build field table
                    table_data = []
                    row = []
                    for j, field in enumerate(section.fields):
                        if not field.visible:
                            continue
                        val = edi_doc.header.get(field.key, "—")
                        if field.type == "currency":
                            try:
                                val = f"${float(val):,.2f}"
                            except:
                                pass
                        
                        row.append([Paragraph(field.label, label_style), 
                                   Paragraph(str(val), value_style)])
                        
                        if len(row) == 2:
                            table_data.append([row[0], row[1]])
                            row = []
                    
                    if row:  # Handle odd number of fields
                        table_data.append([row[0], ['', '']])
                    
                    if table_data:
                        # Flatten for table
                        flat_data = []
                        for data_row in table_data:
                            flat_row = []
                            for cell in data_row:
                                if isinstance(cell, list):
                                    flat_row.extend(cell)
                                else:
                                    flat_row.append(cell)
                            flat_data.append(flat_row)
                        
                        t = Table(flat_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.8*inch])
                        t.setStyle(TableStyle([
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                        ]))
                        elements.append(t)
                
                elif section.type == "table":
                    # Build data table - check both doc attribute and doc.header
                    items = getattr(edi_doc, section.data_source_key, None)
                    if items is None and hasattr(edi_doc, 'header') and isinstance(edi_doc.header, dict):
                        items = edi_doc.header.get(section.data_source_key, [])
                    items = items or []
                    if items:
                        headers = [Paragraph(col.label, label_style) for col in section.columns if col.visible]
                        table_data = [headers]
                        
                        for item in items:
                            row_data = []
                            for col in section.columns:
                                if not col.visible:
                                    continue
                                val = item.get(col.key, "—") if isinstance(item, dict) else "—"
                                if col.type == "currency":
                                    try:
                                        val = f"${float(val):,.2f}"
                                    except:
                                        pass
                                row_data.append(Paragraph(str(val), value_style))
                            table_data.append(row_data)
                        
                        t = Table(table_data)
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#64748b')),
                            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                            ('FONTSIZE', (0, 0), (-1, -1), 9),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('TOPPADDING', (0, 0), (-1, 0), 8),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ]))
                        elements.append(t)
                
                elements.append(Spacer(1, 8))
        
        doc_pdf.build(elements)
        return buffer.getvalue()

    def generate_excel(self, documents: list) -> bytes:
        """
        Generate Excel from documents using layout config.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for i, edi_doc in enumerate(documents):
            ref = edi_doc.header.get('po_number') or edi_doc.header.get('credit_debit_number', f'Doc{i+1}')
            sheet_name = f"{edi_doc.transaction_type}_{str(ref)[:20]}"[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            
            current_row = 1
            
            # Title - support custom placeholders from header
            try:
                format_data = {
                    'name': edi_doc.transaction_name,
                    'ref_number': ref,
                    **edi_doc.header
                }
                title = self.config.title_format.format(**format_data)
            except KeyError:
                title = f"{edi_doc.transaction_name}"
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="1E40AF")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 2
            
            # Render sections
            for section in self.config.sections:
                if not section.visible:
                    continue
                
                # Section title
                ws.cell(row=current_row, column=1, value=section.title.upper())
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=10, color="64748B")
                current_row += 1
                
                if section.type == "fields":
                    col = 1
                    for field in section.fields:
                        if not field.visible:
                            continue
                        val = edi_doc.header.get(field.key, "—")
                        if field.type == "currency":
                            try:
                                val = float(val)
                            except:
                                pass
                        
                        ws.cell(row=current_row, column=col, value=field.label)
                        ws.cell(row=current_row, column=col).font = Font(bold=True, size=9, color="64748B")
                        ws.cell(row=current_row, column=col+1, value=val)
                        
                        col += 2
                        if col > 4:
                            col = 1
                            current_row += 1
                    
                    current_row += 2
                
                elif section.type == "table":
                    # Get table data - check both doc attribute and doc.header
                    items = getattr(edi_doc, section.data_source_key, None)
                    if items is None and hasattr(edi_doc, 'header') and isinstance(edi_doc.header, dict):
                        items = edi_doc.header.get(section.data_source_key, [])
                    items = items or []
                    if items:
                        # Headers
                        visible_cols = [col for col in section.columns if col.visible]
                        for j, col in enumerate(visible_cols, 1):
                            cell = ws.cell(row=current_row, column=j, value=col.label)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='left')
                        current_row += 1
                        
                        # Data rows
                        for item in items:
                            for j, col in enumerate(visible_cols, 1):
                                val = item.get(col.key, "—") if isinstance(item, dict) else "—"
                                if col.type == "currency":
                                    try:
                                        val = float(val)
                                    except:
                                        pass
                                cell = ws.cell(row=current_row, column=j, value=val)
                                cell.border = thin_border
                            current_row += 1
                    
                    current_row += 1
            
            # Auto-adjust column widths
            for column in range(1, 10):
                ws.column_dimensions[get_column_letter(column)].width = 18
        
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def render_content(self, doc) -> str:
        """
        Renders the sections for a single document based on config.
        Returns HTML string (without <html> body wrappers).
        """
        sections_html = []
        for section in self.config.sections:
            if not section.visible:
                continue
            
            content = ""
            if section.type == "fields":
                content = self._render_fields_section(section, doc.header)
            elif section.type == "table":
                content = self._render_table_section(section, doc)
            elif section.type == "grid":
                 content = self._render_placeholder(section)
            
            if content:
                 sections_html.append(f"""
                    <div class="section">
                        <div class="section-title">{section.title}</div>
                        {content}
                    </div>
                 """)
        return "".join(sections_html)

    def _wrap_in_html(self, title: str, body_content: str) -> bytes:
        css = self._get_premium_css()
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>{css}</style>
        </head>
        <body>
            {body_content}
        </body>
        </html>
        """.encode('utf-8')

    def _render_fields_section(self, section: LayoutSection, data: Dict[str, Any]) -> str:
        fields_html = []
        for field in section.fields:
            if not field.visible:
                continue
                
            val = data.get(field.key, "—")
            
            # Format value
            if field.type == "currency":
                try: val = f"${float(val):,.2f}"
                except: pass
            
            style_class = f"style-{field.style}" if field.style else ""
            
            fields_html.append(f"""
                <div class="field-group">
                    <div class="label">{field.label}</div>
                    <div class="value {style_class}">{val}</div>
                </div>
            """)
            
        return f'<div class="fields-grid">{"".join(fields_html)}</div>'

    def _render_table_section(self, section: LayoutSection, doc) -> str:
        # Get list source - check both doc attribute and doc.header
        items = getattr(doc, section.data_source_key, None)
        if items is None and hasattr(doc, 'header') and isinstance(doc.header, dict):
            items = doc.header.get(section.data_source_key, [])
        items = items or []
        
        headers = "".join([f'<th width="{col.width or ""}">{col.label}</th>' for col in section.columns if col.visible])
        
        rows = []
        for item in items:
            cells = []
            for col in section.columns:
                if not col.visible:
                    continue
                val = item.get(col.key, "—") if isinstance(item, dict) else "—"
                # Format
                if col.type == "currency":
                    try: val = f"${float(val):,.2f}"
                    except: pass
                
                style = f'class="font-{col.style}"' if col.style else ""
                cells.append(f'<td {style}>{val}</td>')
            
            rows.append(f'<tr>{"".join(cells)}</tr>')
            
        return f"""
        <div style="overflow-x: auto;">
            <table>
                <thead><tr>{headers}</tr></thead>
                <tbody>{"".join(rows)}</tbody>
            </table>
        </div>
        """

    def _render_placeholder(self, section):
        return f'<div style="padding:10px; color:#666;">Legacy Section: {section.title} (Not yet dynamic)</div>'

    def _get_premium_css(self) -> str:
        return """
            body { font-family: 'Inter', sans-serif; background: #f8fafc; color: #1e293b; margin: 0; padding: 20px; }
            .main-container { max_width: 1000px; margin: 0 auto; background: white; border-radius: 12px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); overflow: hidden; }
            .nav-card { background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%); padding: 24px; color: white; display: flex; justify-content: space-between; align-items: center; }
            .nav-card h1 { margin: 0; font-size: 20px; font-weight: 600; }
            .section { padding: 24px; border-bottom: 1px solid #e2e8f0; }
            .section-title { font-size: 14px; text-transform: uppercase; letter-spacing: 0.05em; color: #64748b; font-weight: 600; margin-bottom: 16px; }
            
            .fields-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 24px; }
            .label { font-size: 12px; color: #64748b; margin-bottom: 4px; }
            .value { font-size: 14px; font-weight: 500; color: #0f172a; }
            .style-bold { font-weight: 700; }
            .style-highlight { color: #2563eb; font-weight: 600; }
            
            table { width: 100%; border-collapse: collapse; font-size: 14px; }
            th { text-align: left; padding: 12px; background: #f8fafc; color: #64748b; font-weight: 600; border-bottom: 2px solid #e2e8f0; }
            td { padding: 12px; border-bottom: 1px solid #f1f5f9; vertical-align: top; }
            tr:last-child td { border-bottom: none; }
            .font-bold { font-weight: 600; }
        """

