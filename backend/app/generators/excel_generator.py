"""
Excel Generator for EDI documents.
"""

from typing import Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.parsers.base import EDIDocument


class ExcelGenerator:
    """Generate Excel output from parsed EDI documents."""
    
    # Styling constants
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def generate(self, document: EDIDocument, output_path: Optional[str] = None) -> bytes:
        """
        Generate an Excel file from an EDI document.
        
        Args:
            document: Parsed EDI document
            output_path: Optional path to save the Excel file
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Create worksheets
        self._create_header_sheet(wb, document)
        self._create_line_items_sheet(wb, document)
        self._create_summary_sheet(wb, document)
        
        # Remove default sheet if we created others
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        excel_bytes = output.read()
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(excel_bytes)
        
        return excel_bytes
    
    def generate_all(self, documents: list, output_path: Optional[str] = None) -> bytes:
        """
        Generate an Excel file from multiple EDI documents.
        Creates separate sheets for each PO.
        
        Args:
            documents: List of parsed EDI documents
            output_path: Optional path to save the Excel file
            
        Returns:
            Excel file as bytes
        """
        if not documents:
            return self.generate(EDIDocument(transaction_type="850", transaction_name="Purchase Order"))
        
        if len(documents) == 1:
            return self.generate(documents[0], output_path)
        
        wb = Workbook()
        
        # Create a summary sheet first
        ws_summary = wb.active
        ws_summary.title = "All POs Summary"
        
        ws_summary["A1"] = f"EDI Conversion - {len(documents)} Purchase Orders"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1E40AF")
        ws_summary.merge_cells("A1:E1")
        
        # Headers for summary
        headers = ["PO Number", "PO Date", "Line Items", "Total Amount", "Ship To"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=3, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Add each document to summary
        for row_idx, doc in enumerate(documents, 4):
            po_num = doc.header.get("po_number", "—")
            po_date = doc.header.get("po_date", "—")
            line_count = len(doc.line_items)
            total = doc.summary.get("total_amount") or doc.summary.get("calculated_total") or "—"
            if isinstance(total, (int, float)):
                total = f"${total:,.2f}"
            
            # Get ship to name
            ship_to = "—"
            for party in doc.header.get("parties", []):
                if party.get("type_code") == "ST" or "ship" in party.get("type", "").lower():
                    ship_to = party.get("name", "—")
                    break
            
            ws_summary.cell(row=row_idx, column=1, value=po_num).border = self.THIN_BORDER
            ws_summary.cell(row=row_idx, column=2, value=po_date).border = self.THIN_BORDER
            ws_summary.cell(row=row_idx, column=3, value=line_count).border = self.THIN_BORDER
            ws_summary.cell(row=row_idx, column=4, value=total).border = self.THIN_BORDER
            ws_summary.cell(row=row_idx, column=5, value=ship_to).border = self.THIN_BORDER
        
        # Column widths for summary
        ws_summary.column_dimensions["A"].width = 18
        ws_summary.column_dimensions["B"].width = 14
        ws_summary.column_dimensions["C"].width = 12
        ws_summary.column_dimensions["D"].width = 16
        ws_summary.column_dimensions["E"].width = 30
        
        # Create sheets for each document
        for idx, doc in enumerate(documents, 1):
            po_num = doc.header.get("po_number", f"PO_{idx}")
            # Sheet names are limited to 31 chars
            sheet_name = f"PO {idx} - {po_num}"[:31]
            
            ws = wb.create_sheet(sheet_name)
            if doc.transaction_type == "812":
                 self._populate_812_sheet(ws, doc, idx, len(documents))
            else:
                 self._populate_document_sheet(ws, doc, idx, len(documents))
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        excel_bytes = output.read()
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(excel_bytes)
        
        return excel_bytes
    
    def _populate_document_sheet(self, ws, document: EDIDocument, idx: int, total: int) -> None:
        """Populate a worksheet with a single document's data."""
        po_num = document.header.get("po_number", "—")
        
        # Title
        ws["A1"] = f"Purchase Order {idx} of {total} — PO #{po_num}"
        ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells("A1:F1")
        
        # Document info section
        ws["A3"] = "Order Information"
        ws["A3"].font = Font(bold=True, size=12)
        
        info_data = [
            ("PO Number", document.header.get("po_number", "—")),
            ("PO Date", document.header.get("po_date", "—")),
            ("Purpose", document.header.get("purpose", "—")),
            ("Order Type", document.header.get("order_type", "—")),
            ("Sender ID", document.sender_id or "—"),
            ("Receiver ID", document.receiver_id or "—"),
            ("Payment Terms", document.header.get("payment_terms", "—")),
            ("F.O.B.", document.header.get("fob", "—")),
        ]
        
        row = 4
        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, color="64748B")
            ws[f"B{row}"] = str(value) if value else "—"
            row += 1
        
        # Line Items section
        row += 1
        ws[f"A{row}"] = f"Line Items ({len(document.line_items)} items)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        if document.line_items:
            # Headers
            headers = ["Line", "Product ID", "Description", "Qty", "Unit", "Unit Price", "Total"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
            row += 1
            
            # Data
            for item in document.line_items:
                unit_price = item.get("unit_price", "")
                total_val = item.get("total", "")
                
                try:
                    unit_price = f"${float(unit_price):,.2f}" if unit_price else "—"
                except (ValueError, TypeError):
                    pass
                
                try:
                    total_val = f"${float(total_val):,.2f}" if total_val else "—"
                except (ValueError, TypeError):
                    pass
                
                ws.cell(row=row, column=1, value=item.get("line_number", "—")).border = self.THIN_BORDER
                ws.cell(row=row, column=2, value=item.get("product_id", "—")).border = self.THIN_BORDER
                ws.cell(row=row, column=3, value=item.get("description", "—")).border = self.THIN_BORDER
                ws.cell(row=row, column=4, value=item.get("quantity", "—")).border = self.THIN_BORDER
                ws.cell(row=row, column=5, value=item.get("unit", "—")).border = self.THIN_BORDER
                ws.cell(row=row, column=6, value=unit_price).border = self.THIN_BORDER
                ws.cell(row=row, column=7, value=total_val).border = self.THIN_BORDER
                row += 1
        
        # Summary
        row += 1
        if document.summary:
            ws[f"A{row}"] = "Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            total_amount = document.summary.get("total_amount") or document.summary.get("calculated_total")
            if total_amount:
                ws[f"A{row}"] = "Total Amount"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = f"${total_amount:,.2f}"
                ws[f"B{row}"].font = Font(bold=True, color="059669", size=14)
        
        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
    
    def _create_header_sheet(self, wb: Workbook, document: EDIDocument) -> None:
        """Create the header/document info sheet."""
        ws = wb.active
        ws.title = "Document Info"
        
        # Title
        ws["A1"] = f"EDI {document.transaction_type} - {document.transaction_name}"
        ws["A1"].font = Font(bold=True, size=16, color="1E40AF")
        ws.merge_cells("A1:B1")
        
        # Document info
        info_data = [
            ("Transaction Type", document.transaction_type),
            ("Transaction Name", document.transaction_name),
            ("Sender ID", document.sender_id or "—"),
            ("Receiver ID", document.receiver_id or "—"),
            ("Control Number", document.control_number or "—"),
            ("Date", document.date or "—"),
        ]
        
        # Add header info
        for key, value in document.header.items():
            if not isinstance(value, (list, dict)):
                info_data.append((key.replace("_", " ").title(), str(value)))
        
        row = 3
        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1
        
        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
    
    def _create_line_items_sheet(self, wb: Workbook, document: EDIDocument) -> None:
        """Create the line items sheet."""
        if not document.line_items:
            return
        
        ws = wb.create_sheet("Line Items")
        
        # Determine columns from first item, excluding dict/list values
        if document.line_items:
            first_item = document.line_items[0]
            columns = [k for k, v in first_item.items() 
                      if not isinstance(v, (dict, list))]
        else:
            return
        
        # Headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Data rows
        for row_idx, item in enumerate(document.line_items, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = item.get(col_name, "")
                # Convert to string if not a basic type
                if isinstance(value, (dict, list)):
                    value = str(value)
                elif value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        # Auto-fit columns
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    def _create_summary_sheet(self, wb: Workbook, document: EDIDocument) -> None:
        """Create the summary sheet."""
        if not document.summary:
            return
        
        ws = wb.create_sheet("Summary")
        
        ws["A1"] = "Summary"
        ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
        
        row = 3
        for key, value in document.summary.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = str(value)
            row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
    def _populate_812_sheet(self, ws, document: EDIDocument, idx: int, total: int) -> None:
        """Populate a worksheet with 812 Credit/Debit Adjustment data."""
        po_num = document.header.get("credit_debit_number", "—")
        
        # Title
        ws["A1"] = f"Credit/Debit Adjustment {idx} of {total} — Memo #{po_num}"
        ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells("A1:F1")
        
        # Document info section
        ws["A3"] = "Adjustment Information"
        ws["A3"].font = Font(bold=True, size=12)
        
        info_data = [
            ("Date", document.header.get("adjustment_date", "—")),
            ("Adj Number", document.header.get("credit_debit_number", "—")),
            ("Handling Code", document.header.get("transaction_handling_desc", "—")),
            ("Amount", document.header.get("amount", "—")),
            ("Flag Code", document.header.get("credit_debit_flag_desc", "—")),
            ("Invoice #", document.header.get("invoice_number", "—")),
            ("PO Number", document.header.get("po_number", "—")),
            ("Purpose", document.header.get("purpose", "—")),
            ("Type", document.header.get("transaction_type_desc", "—")),
            ("Currency", document.header.get("currency", "—")),
        ]
        
        row = 4
        # Format Amount in header
        formatted_info = []
        for label, val in info_data:
            if label == "Amount" and val != "—":
                try: val = f"${float(val):,.2f}"
                except: pass
            formatted_info.append((label, val))

        for label, value in formatted_info:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, color="64748B")
            ws[f"B{row}"] = str(value)
            row += 1
            
        # Contacts
        contacts = document.header.get("contacts", [])
        if contacts:
            row += 1
            ws[f"A{row}"] = "Contacts"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for c in contacts:
                if c.get("name"):
                    ws[f"A{row}"] = "Name"
                    ws[f"B{row}"] = c["name"]
                    row += 1
                if c.get("comm_number"):
                    ws[f"A{row}"] = "Tel"
                    ws[f"B{row}"] = c["comm_number"]
                    row += 1

        # Parties
        row += 1
        ws[f"A{row}"] = "Parties"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        for party in document.header.get("parties", []):
            name = party.get("name", "—")
            ptype = party.get("type", "")
            ws[f"A{row}"] = ptype
            ws[f"A{row}"].font = Font(bold=True, color="64748B")
            ws[f"B{row}"] = name
            row += 1

        # Adjustment Details section
        row += 1
        ws[f"A{row}"] = f"Adjustment Details ({len(document.line_items)} items)"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        if document.line_items:
            # Headers
            headers = ["Reason", "Flag", "Amount", "Quantity", "Unit Price", "Unit", "Parts", "Message"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.THIN_BORDER
            row += 1
            
            # Data
            for item in document.line_items:
                # Format money fields
                amt = item.get("adjustment_amount", "")
                try: amt = f"${float(amt):,.2f}" if amt else ""
                except: pass

                price = item.get("unit_price", "")
                try: price = f"${float(price):.2f}" if price else ""
                except: pass
                
                # Format parts
                parts = item.get("part_numbers", {})
                parts_str = ", ".join([f"{k}:{v}" for k,v in parts.items()])

                ws.cell(row=row, column=1, value=item.get("adjustment_reason", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=2, value=item.get("credit_debit_type", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=3, value=amt).border = self.THIN_BORDER
                ws.cell(row=row, column=4, value=item.get("quantity", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=5, value=price).border = self.THIN_BORDER
                ws.cell(row=row, column=6, value=item.get("unit", "")).border = self.THIN_BORDER
                ws.cell(row=row, column=7, value=parts_str).border = self.THIN_BORDER
                ws.cell(row=row, column=8, value=item.get("message", "")).border = self.THIN_BORDER
                row += 1
        
        # Summary
        row += 1
        if document.summary:
            ws[f"A{row}"] = "Summary"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            row += 1
            
            total_amount = document.summary.get("total_amount") or document.summary.get("calculated_total")
            if total_amount:
                ws[f"A{row}"] = "Total Amount"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = f"${total_amount:,.2f}"
                ws[f"B{row}"].font = Font(bold=True, color="059669", size=14)
        
        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 40
